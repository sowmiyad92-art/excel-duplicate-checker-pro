from flask import Flask, render_template, request, session, send_file, redirect, url_for, flash, jsonify
import os
import pandas as pd
from werkzeug.utils import secure_filename
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import logging

app = Flask(__name__)
app.secret_key = 'your_updated_secret_key_change_this'
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB limit

# Create directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_info(filepath):
    """Get basic file information"""
    try:
        file_size = os.path.getsize(filepath)
        return {
            'size': f"{file_size / (1024*1024):.2f} MB",
            'modified': datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d %H:%M'),
            'name': os.path.basename(filepath)
        }
    except:
        return {'size': 'Unknown', 'modified': 'Unknown', 'name': 'Unknown'}

def analyze_data_quality(df):
    """Enhanced data quality analysis"""
    analysis = {
        'total_rows': len(df),
        'total_columns': len(df.columns),
        'memory_usage': f"{df.memory_usage(deep=True).sum() / 1024:.2f} KB",
        'blank_cells': int(df.isnull().sum().sum()),
        'blank_percentage': round((df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100, 2) if len(df) > 0 else 0,
        'data_types': df.dtypes.value_counts().to_dict(),
        'column_stats': {}
    }

    # Analyze each column
    for col in df.columns:
        col_stats = {
            'dtype': str(df[col].dtype),
            'non_null_count': int(df[col].count()),
            'null_count': int(df[col].isnull().sum()),
            'unique_values': int(df[col].nunique()),
            'duplicates_in_column': int(df[col].duplicated().sum())
        }

        # Add sample values for object columns
        if df[col].dtype == 'object' and not df[col].empty:
            try:
                sample_values = df[col].dropna().astype(str).unique()[:3].tolist()
                col_stats['sample_values'] = sample_values
            except:
                col_stats['sample_values'] = []

        analysis['column_stats'][col] = col_stats

    return analysis

def highlight_duplicates(df, subset, duplicate_type='all'):
    """Enhanced duplicate highlighting with better styling"""
    df_copy = df.copy()

    if duplicate_type == 'all':
        dup_mask = df_copy.duplicated(subset=subset, keep=False)
    elif duplicate_type == 'except_first':
        dup_mask = df_copy.duplicated(subset=subset, keep='first')
    elif duplicate_type == 'except_last':
        dup_mask = df_copy.duplicated(subset=subset, keep='last')
    else:
        dup_mask = df_copy.duplicated(subset=subset, keep=False)

    def highlight_row(row):
        if dup_mask[row.name]:
            return ['background-color: #fff3cd; border-left: 3px solid #ffc107; font-weight: bold;' for _ in row]
        else:
            return ['background-color: #f8f9fa;' for _ in row]

    styled = df_copy.style.apply(highlight_row, axis=1)
    styled = styled.set_table_styles([
        {'selector': 'th', 'props': [('background-color', '#007bff'), ('color', 'white'), ('font-weight', 'bold')]},
        {'selector': 'td', 'props': [('padding', '8px'), ('border', '1px solid #dee2e6')]},
        {'selector': 'table', 'props': [('border-collapse', 'collapse'), ('width', '100%')]}
    ])

    return styled.to_html()

def normalize_data_for_comparison(df, columns):
    """Normalize data for better duplicate detection"""
    df_normalized = df.copy()

    for col in columns:
        if col in df_normalized.columns:
            # Handle string columns
            if df_normalized[col].dtype == 'object':
                # Convert to string, strip whitespace, convert to lowercase
                df_normalized[col] = df_normalized[col].astype(str).str.strip().str.lower()
                # Replace multiple spaces with single space
                df_normalized[col] = df_normalized[col].str.replace(r'\s+', ' ', regex=True)
                # Handle common variations
                df_normalized[col] = df_normalized[col].replace({
                    'nan': pd.NA,
                    'none': pd.NA,
                    '': pd.NA,
                    'null': pd.NA
                })

            # Handle numeric columns
            elif df_normalized[col].dtype in ['float64', 'float32']:
                # Round to 6 decimal places to handle floating point precision issues
                df_normalized[col] = df_normalized[col].round(6)

    return df_normalized

def create_enhanced_reports(original_df, duplicates_df, cleaned_df, selected_columns, analysis, duplicate_type):
    """Create comprehensive Excel reports with formatting"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # File paths
    original_path = os.path.join(RESULT_FOLDER, f"Original_Data_{timestamp}.xlsx")
    duplicates_path = os.path.join(RESULT_FOLDER, f"Duplicates_Only_{timestamp}.xlsx")
    cleaned_path = os.path.join(RESULT_FOLDER, f"Cleaned_Data_{timestamp}.xlsx")
    summary_path = os.path.join(RESULT_FOLDER, f"Analysis_Summary_{timestamp}.xlsx")

    # Create summary report
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_data = {
            'Metric': [
                'Analysis Date',
                'Total Rows',
                'Total Columns',
                'Duplicate Rows Found',
                'Cleaned Rows Remaining',
                'Removal Rate (%)',
                'Blank Cells',
                'Blank Percentage (%)',
                'Analysis Columns',
                'Duplicate Detection Type',
                'Memory Usage'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                analysis['total_rows'],
                analysis['total_columns'],
                len(duplicates_df),
                len(cleaned_df),
                round((len(duplicates_df) / analysis['total_rows'] * 100), 2) if analysis['total_rows'] > 0 else 0,
                analysis['blank_cells'],
                analysis['blank_percentage'],
                ', '.join(selected_columns),
                duplicate_type.replace('_', ' ').title(),
                analysis['memory_usage']
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # Column analysis
        col_analysis = []
        for col, stats in analysis['column_stats'].items():
            col_analysis.append({
                'Column Name': col,
                'Data Type': stats['dtype'],
                'Non-Null Count': stats['non_null_count'],
                'Null Count': stats['null_count'],
                'Unique Values': stats['unique_values'],
                'Column Duplicates': stats['duplicates_in_column'],
                'Sample Values': ', '.join(stats.get('sample_values', []))[:50] + '...' if len(', '.join(stats.get('sample_values', []))) > 50 else ', '.join(stats.get('sample_values', []))
            })
        pd.DataFrame(col_analysis).to_excel(writer, sheet_name="Column Analysis", index=False)

    # Save individual files
    with pd.ExcelWriter(original_path, engine="openpyxl") as writer:
        original_df.to_excel(writer, sheet_name="Original Data", index=False)

    if not duplicates_df.empty:
        with pd.ExcelWriter(duplicates_path, engine="openpyxl") as writer:
            duplicates_df.to_excel(writer, sheet_name="Duplicate Rows", index=False)

        # Format duplicates file with highlighting
        wb = load_workbook(duplicates_path)
        ws = wb["Duplicate Rows"]
        yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Format data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = yellow_fill

        wb.save(duplicates_path)

    with pd.ExcelWriter(cleaned_path, engine="openpyxl") as writer:
        cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)

    # Create ZIP file
    zip_path = os.path.join(RESULT_FOLDER, f"Duplicate_Analysis_Report_{timestamp}.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(summary_path, f"Analysis_Summary_{timestamp}.xlsx")
        zipf.write(original_path, f"Original_Data_{timestamp}.xlsx")
        if not duplicates_df.empty:
            zipf.write(duplicates_path, f"Duplicates_Only_{timestamp}.xlsx")
        zipf.write(cleaned_path, f"Cleaned_Data_{timestamp}.xlsx")

    return zip_path

@app.route("/", methods=["GET", "POST"])
def index():
    sheet_names = []
    columns = []
    selected_sheet = None
    selected_columns = []
    duplicates_html = None
    sheet_preview_html = None
    data_stats = None
    analysis = None
    error_msg = None
    success_msg = None
    file_uploaded = False
    file_info = None
    duplicate_type = 'all'

    filepath = session.get("file_path")

    if request.method == "POST":
        uploaded_file = request.files.get("file1")
        selected_sheet = request.form.get("sheet")
        selected_columns = request.form.getlist("selected_columns")
        duplicate_type = request.form.get("duplicate_type", "all")

        # Handle file upload
        if uploaded_file and uploaded_file.filename != "":
            if not allowed_file(uploaded_file.filename):
                error_msg = "Please upload a valid Excel file (.xlsx, .xls) or CSV file."
                return render_template("index.html", error_msg=error_msg)

            # Check file size
            uploaded_file.seek(0, os.SEEK_END)
            file_size = uploaded_file.tell()
            uploaded_file.seek(0)

            if file_size > MAX_FILE_SIZE:
                error_msg = f"File too large. Maximum size is {MAX_FILE_SIZE/(1024*1024):.0f}MB."
                return render_template("index.html", error_msg=error_msg)

            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
            uploaded_file.save(filepath)
            session["file_path"] = filepath
            file_uploaded = True
            success_msg = f"File '{filename}' uploaded successfully!"

        # Check if file exists
        if not filepath or not os.path.exists(filepath):
            error_msg = "Please upload an Excel file first."
            return render_template("index.html", error_msg=error_msg, file_uploaded=file_uploaded)

        try:
            file_info = get_file_info(filepath)

            # Read file based on extension
            if filepath.lower().endswith('.csv'):
                # Handle CSV files
                try:
                    df = pd.read_csv(filepath, encoding='utf-8')
                    sheet_names = ['Sheet1']
                    selected_sheet = 'Sheet1'
                except UnicodeDecodeError:
                    try:
                        df = pd.read_csv(filepath, encoding='latin-1')
                        sheet_names = ['Sheet1']
                        selected_sheet = 'Sheet1'
                    except Exception as e:
                        error_msg = f"Error reading CSV file: {str(e)}"
                        return render_template("index.html", error_msg=error_msg, file_uploaded=file_uploaded)
            else:
                # Handle Excel files
                sheet_data = pd.read_excel(filepath, sheet_name=None)
                sheet_names = list(sheet_data.keys())

                if not selected_sheet or selected_sheet not in sheet_names:
                    selected_sheet = sheet_names[0]

                df = sheet_data[selected_sheet]

            # Data validation
            if df.empty:
                error_msg = "The selected sheet/file is empty."
                return render_template("index.html", 
                                     error_msg=error_msg, 
                                     file_uploaded=file_uploaded, 
                                     file_info=file_info,
                                     sheet_names=sheet_names,
                                     selected_sheet=selected_sheet)

            # Clean column names
            df.columns = df.columns.astype(str).str.strip()
            columns = [col for col in df.columns if col and str(col) != 'nan']

            # Perform data analysis
            analysis = analyze_data_quality(df)

            # Process column selection for duplicate checking
            valid_columns = [col for col in selected_columns if col in df.columns]

            if valid_columns:
                try:
                    # Normalize data for better duplicate detection
                    df_normalized = normalize_data_for_comparison(df, valid_columns)

                    # Find duplicates based on type
                    if duplicate_type == 'all':
                        mask = df_normalized.duplicated(subset=valid_columns, keep=False)
                    elif duplicate_type == 'except_first':
                        mask = df_normalized.duplicated(subset=valid_columns, keep='first')
                    elif duplicate_type == 'except_last':
                        mask = df_normalized.duplicated(subset=valid_columns, keep='last')
                    else:
                        mask = df_normalized.duplicated(subset=valid_columns, keep=False)

                    duplicates = df[mask].copy()  # Use original data for display
                    cleaned_df = df[~mask].copy()

                    # Generate HTML for display
                    if not duplicates.empty:
                        duplicates_html = duplicates.head(100).to_html(
                            classes="table table-striped table-hover table-sm", 
                            index=False,
                            table_id="duplicatesTable"
                        )
                        success_msg = f"Analysis complete! Found {len(duplicates)} duplicate rows."
                    else:
                        success_msg = "Great! No duplicates found with the selected criteria."

                    # Generate preview with highlighting
                    sheet_preview_html = highlight_duplicates(df.head(20), valid_columns, duplicate_type)

                    # Update data stats
                    data_stats = {
                        "total_rows": len(df),
                        "total_columns": len(df.columns),
                        "blank_cells": analysis['blank_cells'],
                        "duplicate_rows": len(duplicates),
                        "cleaned_rows": len(cleaned_df),
                        "removal_percentage": round((len(duplicates) / len(df) * 100), 2) if len(df) > 0 else 0
                    }

                    # Save enhanced reports
                    zip_path = create_enhanced_reports(df, duplicates, cleaned_df, valid_columns, analysis, duplicate_type)
                    session["download_path"] = zip_path

                except Exception as e:
                    error_msg = f"Error during duplicate analysis: {str(e)}"
                    logger.error(f"Duplicate analysis error: {e}")
            else:
                # Show preview without analysis
                sheet_preview_html = df.head(20).style.set_table_attributes('class="table table-striped table-sm"').to_html()
                if selected_columns:
                    error_msg = "Selected columns not found in the data. Please check column names."

        except Exception as e:
            error_msg = f"Error processing the file: {str(e)}"
            logger.error(f"File processing error: {e}")

    return render_template("index.html",
                           sheet_names=sheet_names,
                           selected_sheet=selected_sheet,
                           columns=columns,
                           selected_columns=selected_columns,
                           duplicates_html=duplicates_html,
                           sheet_preview_html=sheet_preview_html,
                           data_stats=data_stats,
                           analysis=analysis,
                           error_msg=error_msg,
                           success_msg=success_msg,
                           file_uploaded=file_uploaded,
                           file_info=file_info,
                           duplicate_type=duplicate_type,
                           current_sheet_name=selected_sheet)

@app.route("/download")
def download():
    download_path = session.get("download_path")
    if download_path and os.path.exists(download_path):
        return send_file(download_path, as_attachment=True)
    else:
        flash("No report available for download. Please analyze data first.", "warning")
        return redirect(url_for("index"))

@app.route("/reset")
def reset():
    # Clean up files
    try:
        if "file_path" in session:
            filepath = session["file_path"]
            if os.path.exists(filepath):
                os.remove(filepath)

        if "download_path" in session:
            download_path = session["download_path"]
            if os.path.exists(download_path):
                os.remove(download_path)
    except Exception as e:
        logger.error(f"Error cleaning up files: {e}")

    session.clear()
    flash("Session reset successfully!", "success")
    return redirect(url_for("index"))

@app.route("/api/preview")
def api_preview():
    """API endpoint for live preview of selected columns"""
    try:
        filepath = session.get("file_path")
        if not filepath or not os.path.exists(filepath):
            return jsonify({"error": "No file uploaded"}), 400

        sheet = request.args.get("sheet", "Sheet1")
        columns = request.args.getlist("columns[]")

        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath, sheet_name=sheet)

        if columns:
            valid_columns = [col for col in columns if col in df.columns]
            if valid_columns:
                preview_data = df[valid_columns].head(5).fillna('').to_dict('records')
                return jsonify({"preview": preview_data, "columns": valid_columns})

        return jsonify({"preview": [], "columns": []})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=True)
